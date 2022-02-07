#예외 강제 발생
def ten_div(x):
    if x>10:
        #강제로 예외를 발생시켜서 아래 문장을 수행하지 않음
        raise Exception("숫자가 너무 큽니다")
    return 10/x

#기본적인 예외 처리 구문을 이용해서 예외가 발생하더라도 중지되지 않도록 함
try:
    print(ten_div(2))
    print(ten_div(15)) #예외가 발생해서 프로그램이 중단
#예외가 발생했을 때 예외 메시지 출력
except Exception as e:
    print(e)

else:
    print('예외가 발생하지 않은 경우 수행')

finally:
    print('프로그램 종료')


#일반 함수 호출
import threading, time
def threadEx(id):
    for i in range(10):
        print('id={0} --> {1}'.format(id, i))
        time.sleep(1)

for i in range(2):
    id = ("{0}번 스레드".format(i))
    th = threading.Thread(target=threadEx, args=(id,))  # 쓰레드 함수를 target인수로, 쓰레드로 전달될 인수를 args인수에 전달
    th.start()
print("메인 종료")          # 쓰레드 실행 시작

#singer.csv 읽기
import csv
with open('./data/data/singer.csv', 'r') as f:
    content = f.read()
    data = content.split(',')
    print(type(data))
    for imsi in data:
        print(imsi)


# csv 모듈을 이용해서 읽기 - csv.reader(file객체)를 대입해서 읽음
# 줄 바꿈 별로 하나의 list를 별도로 생성
import csv

with open('./data/data/singer.csv', 'r') as f:
    # CSV로 읽기
    # 구분자가 ,가 아니면 delimiter에 구분자를 설정해주면 됨
    data = csv.reader(f)
    print(type(data))  # 자료형 확인
    # 전체 출력
    for imsi in data:
        print(imsi)


#기록하기
import csv
#newline옵션을 이용해서 빈줄이 생기지 않도록 설정
with open('./data/data/singer.csv', 'a' , newline='') as f:
    #CSV로 읽기
    wr = csv.writer(f)
    wr.writerow([4, 'zkfls', '1991-05-30', 'aespa'])

#바이트 단위로 파일에 읽고 쓰기
with open('./data/test.bin', 'wb') as f:
    f.write("안녕하세요".encode())

#문자열을 바이트 단위로 기록했으므로 읽을 때는 문자열로 변환해서 읽어야 함
with open('./data/test.bin', 'rb') as f:
    byteAr = f.read()
    #print(byteAr)
    print(byteAr.decode())


##Serializable(직렬화)-객체 단위로 파일에 읽고 쓰는 것
# 직렬화를 위해 DTO클래스 생성
class DTO:
    def setNum(self, num):
        self.num = num

    def setName(self, name):
        self.name = name

    def getNum(self):
        return self.num

    def getName(self):
        return self.name

    def toString(self):
        return "{번호:{0} 이름:{1}}".format(self.num, self.name)


# 직렬화 할 인스턴스 생성
dto1 = DTO()
dto1.setNum(1)
dto1.setName("유관순")

dto2 = DTO()
dto2.setNum(2)
dto2.setName("안중근")

li = [dto1, dto2]

import pickle

with open('./data/data.dat', 'wb') as f:
    pickle.dump(li, f)


#객체 직렬화 – 직렬화
import pickle
try:
    with open('./data/test.txt', 'wb') as f:
        pickle.dump(li, f)
        f.close()
    with open('./data/test.txt', 'rb') as f:
        result = pickle.load(f)
        for temp in result:
            print(temp.toString())
        f.close()
except Exception as e:
    print("예외 발생:", e)
finally:
    f.close()


import pickle
with open('./data/data.dat', 'rb') as f:
    result = pickle.load(f)
    for dto in result:
        print(dto.toString())



import zipfile
filelist = ["./test.txt"]
with zipfile.ZipFile('./test.zip', 'w', compression=zipfile.ZIP_BZIP2) as myzip:
    for temp in filelist:
        myzip.write(temp)
zipfile.ZipFile('./test.zip').extractall()

# 엑셀 파일 읽기
import openpyxl

# 엑셀 파일 포인터
wb = None

try:
    # 엑셀 파일에 포인터를 생성
    wb = openpyxl.load_workbook('./data/data/score.xlsx')
    # print(wb) # 경로 맞나 확인하고 주석처리 ㅎㅎ

    # sheet 가져오기
    ws = wb.active  # 현재 활성화된 sheet를 가져옴.
    # ws = wb.get_sheet_by_name("sheet이름")

    # 행단위 접근 - 각 행을 튜플로 접근
    # 읽어오는 개념이라 못고치도록 주는 것 ,,,,,
    for row in ws.rows:
        print(row)
        # 각 행의 인덱스
        print(row[0].value)
        print(row[0].row)

except Exception as e:
    print(e)

finally:
    if wb != None:
        wb.close()




# 엑셀 파일 읽기
import openpyxl
# 엑셀 파일 포인터
wb = None
try:
    # 엑셀 파일에 포인터를 생성
    wb = openpyxl.load_workbook('./data/data/score.xlsx')
    ws = wb.active  # 현재 활성화된 sheet를 가져옴.
    for row in ws.rows:

        # 1행 5열의 합계라고 설정
        if row[0].row == 1:
            ws.cell(row=row[0].row, column=5).value = "합계"
            continue

        # 합계 구하기
        total = row[1].value + row[2].value + row[3].value
        ws.cell(row=row[0].row, column=5).value = total

    # 파일에 기록
    wb.save("./data/score2.xlsx")
except Exception as e:
    print(e)

finally:
    if wb != None:
        wb.close()


# 데이터베이스 연동
###MYSQL접속
#mysql 접속
import pymysql

con = pymysql.connect(host='127.0.0.1', port=3306, user='root', passwd='1025', db='user00', charset='utf8')

print(con)

con.close()

# 데이터삽입
import sys, pymysql

# 연결 객체 변수 생성
con = None

try:
    # 연결
    con = pymysql.connet(host='localhost', port=3306,
                         db='user11', user='root', passwd='1025', charset='utf8')
    print(con)

except Exception as e:
    print(e)
    print(sys.exc_info())
finally:
    # 연결 해제
    if con != None:
        con.close()

        # 데이터삽입
        import sys, pymysql

        # 연결 객체 변수 생성
        con = None

        try:
            # 연결
            con = pymysql.connet(host='localhost', port=3306,
                                 db='user11', user='root', passwd='1025', charset='utf8')
            print(con)

        except Exception as e:
            print(e)
            print(sys.exc_info())
        finally:
            # 연결 해제
            if con != None:
                con.close()

# 데이터삽입
import sys, pymysql

# 연결 객체 변수 생성
con = None

try:
    # 연결
    con = pymysql.connect(host='localhost', port=3306,
                          db='user11', user='root', passwd='1025', charset='utf8')
    print(con)

    # sql 실행 객체 가져오기
    cursor = con.cursor()
    """
    cursor.execute("insert into usertbl values('dj','dajung',0530 ,'seoul','01077717721','20220207')")
    """
    # 파라미터 매핑
    cursor.execute("insert into usertbl values(%s,%s,%s,%s,%s,%s)",
                   ("g", "eajung", 1215, "seoul", "01077717721", "20220207"))
    con.commit()

except Exception as e:
    print(e)
    print(sys.exc_info())
finally:
    # 연결 해제
    if con != None:
        con.close()


#mongodb서버 연결
from pymongo import MongoClient

con = MongoClient('127.0.0.1')

print(con)


#데이터 삽입
from pymongo import MongoClient

con = MongoClient('127.0.0.1')

#사용할 데이터베이스 연결
db = con.mymongo

#컬렉션 설정 - 테이블 생성 또는 연결
collect = db.users

#삽입할 데이터 생성 - dict
doc1={'empno' : '0001', 'ename' : '강'}
doc2={'empno' : '0002', 'ename' : '김'}
doc3={'empno' : '0003', 'ename' : '박'}
doc4={'empno' : '0004', 'ename' : '이'}

#파이선 , 자바스크립트 , 노드는 행을 만드는 방법이 같음 .

collect.insert_many([doc1, doc2, doc3, doc4])


#데이터 조회
from pymongo import MongoClient

con = MongoClient('127.0.0.1')

#사용할 데이터베이스 연결
db = con.mymongo

#컬렉션 설정 - 테이블 생성 또는 연결
collect = db.users

#조회
#result=collect.find()
#print(result)
#Cursor가 나오면 이터레이터? 생각해야함
for r in result:
    print(r)



#웹에서 데이터 가져오기
#기본패키지 이용
import urllib.request

#데이터 읽기
response = urllib.request.urlopen('https://www.daum.net')
#바이트 배열
data = response.read()
print(data)


import urllib.request
#파라미터 인코딩을 위한 모듈
from urllib.parse import quote

keyword = quote('코로나')

#데이터 읽기
response = urllib.request.urlopen("http://search.hani.co.kr/Search?command=query&keyword='keyword'&sort=d&period=all&media=news")

#바이트 배열
data = response.read()
#읽어온 데이터의 인코딩 확인
encoding = response.info().get_content_charset()
#인코딩 설정
html= data.decode('utf8')
print(html)


#requests 모듈 이용하기
import requests

response = requests.get("https://www.naver.com")
#print(response.text)

response = requests.get("http://www.google.com")
print(response.text)


#이미지 가져와서 저장하기
import requests
imgurl = "https://img1.daumcdn.net/thumb/C428x428/?scode=mtistory2&fname=https%3A%2F%2Ftistory3.daumcdn.net%2Ftistory%2F5007010%2Fattach%2Fbdd39d43763245a2a790df254bb0d222"
response = requests.get(imgurl)

with open('./daj.jpg', 'wb') as f:
    img = response.content
    f.write(img)


#JSON 파싱
#카카오 오픈api데이터 가져오기 - 카테고리 검색
import requests
import json
#약국을 찾겠다 PM9
url = 'https://dapi.kakao.com/v2/local/search/category.json?category_group_code=PM9&y=37.57002838826&x=126.97962084516&radius=5000'
headers ={'Authorization': 'KakaoAK 5957e853b495a0aba8c6eefc5489fce8'}

data = requests.get(url, headers=headers)
print(data.text)


#JSON 파싱
#카카오 오픈api데이터 가져오기 - 카테고리 검색
import requests
import json
#약국을 찾겠다 PM9
url = 'https://dapi.kakao.com/v2/local/search/category.json?category_group_code=PM9&y=37.57002838826&x=126.97962084516&radius=5000'
headers ={'Authorization': 'KakaoAK 5957e853b495a0aba8c6eefc5489fce8'}

data = requests.get(url, headers=headers)
#print(data.text)

#json 파싱
result = json.loads(data.text)
#print(type(result))
#<class 'dict'>일때는
documents = result["documents"]
#print(documents)

for temp in documents:
    print(temp['place_name'], ":", temp['address_name'])


#기상청 오늘의 날씨 사이트에서 지점과 체감온도 , 습도, 풍속 가져오기
import requests, bs4
#html확인
response = requests.get( 'https://www.weather.go.kr/weather/observation/currentweather.jsp')
print(response.text)



#기상청 오늘의 날씨 사이트에서 지점과 체감온도 , 습도, 풍속 가져오기
import requests, bs4
#html확인
response = requests.get( 'https://www.weather.go.kr/weather/observation/currentweather.jsp')
#print(response.text)

soup = bs4.BeautifulSoup(response.text, 'html.parser')
#print(soup)
#원하는 정보가 잇는 테이블을 선택
#table = soup.select('#content_weather > table')
table = soup.select('.table_develop3')
print(table)


#기상청 오늘의 날씨 사이트에서 지점과 체감온도 , 습도, 풍속 가져오기
import requests, bs4
#html확인
response = requests.get( 'https://www.weather.go.kr/weather/observation/currentweather.jsp')
#print(response.text)

soup = bs4.BeautifulSoup(response.text, 'html.parser')
#print(soup)
#원하는 정보가 잇는 테이블을 선택
#table = soup.select('#content_weather > table')
table = soup.select('.table_develop3')
for temp in table[0].find_all('tr'):
    #각 줄의 각 칸들을 list로 변환
    tds=list(temp.find_all('td'))
    #list순회
    for td in tds:
        if td.find('a'):
            print(td.find('a').text, ":", tds[7].text, ":", tds[10].text)



#HTML Parsinig
#https://tv.naver.com/r/ 에서 타이틀 가져오기
import requests, bs4

#html 가져오기
response = requests.get('https://tv.naver.com/r/')
html = response.text

#html을 메모리 펼치기
bs = bs4.BeautifulSoup(html, 'html.parser')
#print(bs)

#필요한데이터선택
tags = bs.select('dl > dt > a > tooltip')


#xml파싱
#동아일보 전체 기사 파싱
import requests, bs4
#html확인
response = requests.get('https://rss.donga.com/total.xml')
#print(response.text)
rss=bs4.BeautifulSoup(response.text, 'lxml-xml')
print(rss)

#item 태그 안의 title 추출
items = rss.find_all('item')
print(items)


#xml파싱
#동아일보 전체 기사 파싱
import requests, bs4
#html확인
response = requests.get('https://rss.donga.com/total.xml')
#print(response.text)
rss=bs4.BeautifulSoup(response.text, 'lxml-xml')
#print(rss)

#item 태그 안의 title 추출
items = rss.find_all('item')
#print(items)
for item in items:
    print(item.find('title').getText())


#selenium
#크롬 실행
from selenium import webdriver

#드라이버 경로를 설정하면 크롬이 실행됨
driver = webdriver.Chrome('c:\\lecture\\chromedriver')


#다음 로그인 하기
#크롬실행
from selenium import webdriver

#드라이버 경로를 설정하면 크롬이 실행됨
driver = webdriver.Chrome('c:\\lecture\\chromedriver')

#다음 로그인 페이지로 이동
driver.get('https://logins.daum.net/accounts/signinform.do?url=https%3A%2F%2Fwww.daum.net')

#아이디와 비밀번호 입력
driver.find_element_by_xpath('//*[@id="id"]').send_keys('id')
driver.find_element_by_xpath('//*[@id="inputPwd"]').send_keys('pw')

#로그인 버튼을 찾아서 클릭
driver.find_element_by_xpaty('//*[@id="loginBtn"]').click()

#특정 카페로 이동
driver.get('http://cafe.daum.net/samhak7/')

#유투브스크롤
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

#드라이버 경로를 설정하면 크롬이 실행됨
driver = webdriver.Chrome('c:\\lecture\\chromedriver')

driver.get('https://www.youtube.com/')

body = driver.find_element_by_tag_name('body')

time.sleep(10)

body.send_keys(Keys.PAGE_DOWN)